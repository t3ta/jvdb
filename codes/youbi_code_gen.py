import openpyxl

workbook = openpyxl.load_workbook('../JV-Data480.xlsx')
sheet = workbook['コード表']
commands = []

for row in range(131, 140):
    values = (
        sheet.cell(row, 3).value or "",
        sheet.cell(row, 4).value or "",
        sheet.cell(row, 5).value or "",
        sheet.cell(row, 6).value or "",
        sheet.cell(row, 7).value or "",
        sheet.cell(row, 8).value or "",
        sheet.cell(row, 9).value or ""
    )

    if values[0] == "":
        continue

    commands.append(
        "INSERT INTO youbi_code VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s');" % values)

    with open("./data/youbi_code_data.sql", mode="w") as f:
        f.write('\n'.join(commands))
