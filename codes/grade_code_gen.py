import openpyxl

workbook = openpyxl.load_workbook('./JV-Data480.xlsx')
sheet = workbook['コード表']
commands = []


for row in range(145, 155):
    values = (
        sheet.cell(row, 3).value or "",
        sheet.cell(row, 4).value or ""
    )
    commands.append("INSERT INTO grade_code VALUES ('%s', '%s');" % values)

    if values[0] == "":
        continue

    with open("./data/grade_code_data.sql", mode="w") as f:
        f.write('\n'.join(commands))
